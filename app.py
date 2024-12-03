# 首先导入和执行 eventlet monkey patch
# import eventlet
# eventlet.monkey_patch()

# 然后导入其他模块
from flask import Flask, render_template, request, send_file, jsonify
from flask_socketio import SocketIO
import os
import logging
from pathlib import Path
from werkzeug.utils import secure_filename
import tempfile
from datetime import datetime
import boto3
from io import BytesIO
import json
from dotenv import load_dotenv
from wifi_processor import process_wifi_data, format_worksheet
import sys
import polars as pl
from openpyxl import Workbook

# 加载 .env 文件（仅在本地开发时需要）
if os.path.exists('.env'):
    load_dotenv()

# 配置日志
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    stream=sys.stdout  # Vercel会捕获stdout的输出
)
logger = logging.getLogger(__name__)

# 创建Flask应用
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 300 * 1024 * 1024  # 300MB
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or os.urandom(24)

# 设置上传目录
UPLOAD_FOLDER = '/tmp/wifi_analysis'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 初始化SocketIO
socketio = SocketIO(app, async_mode='threading', cors_allowed_origins="*")

ALLOWED_EXTENSIONS = {'xlsx'}

# R2配置从环境变量获取
R2_CONFIG = {
    'endpoint_url': os.environ.get('R2_ENDPOINT_URL'),
    'aws_access_key_id': os.environ.get('R2_ACCESS_KEY_ID'),
    'aws_secret_access_key': os.environ.get('R2_SECRET_ACCESS_KEY'),
    'bucket_name': os.environ.get('R2_BUCKET_NAME')
}

# 验证必要的环境变量是否存在
required_env_vars = [
    'R2_ENDPOINT_URL',
    'R2_ACCESS_KEY_ID',
    'R2_SECRET_ACCESS_KEY',
    'R2_BUCKET_NAME'
]

missing_vars = [var for var in required_env_vars if not os.environ.get(var)]
if missing_vars:
    raise RuntimeError(f"Missing required environment variables: {', '.join(missing_vars)}")

# 初始化R2客户端
s3_client = boto3.client(
    's3',
    endpoint_url=R2_CONFIG['endpoint_url'],
    aws_access_key_id=R2_CONFIG['aws_access_key_id'],
    aws_secret_access_key=R2_CONFIG['aws_secret_access_key']
)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def emit_progress(progress, status):
    with app.app_context():
        socketio.emit('progress', {'progress': progress, 'status': status})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return jsonify({'error': '没有选择文件'})

    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({'error': '没有选择文件'})

    try:
        session_id = os.urandom(16).hex()
        total_files = len(files)
        processed_files = 0

        output_filename = 'wifi_statistics_result.xlsx'
        
        # 创建一个内存中的Excel writer
        output_buffer = BytesIO()
        
        # 创建一个字典来存储所有处理后的数据框
        all_dataframes = {}
        
        for file in files:
            if not allowed_file(file.filename):
                continue

            filename = secure_filename(file.filename)
            
            with app.app_context():
                emit_progress(
                    int((processed_files * 100) / total_files),
                    f'正在处理: {filename}'
                )

            try:
                file_content = BytesIO(file.read())
                result_df = process_wifi_data(file_content)
                
                sheet_name = os.path.splitext(filename)[0][:31]

                if result_df is not None and not result_df.is_empty():
                    all_dataframes[sheet_name] = result_df
                    with app.app_context():
                        emit_progress(
                            int(((processed_files + 1) * 100) / total_files),
                            f'已完成处理: {filename}'
                        )
                else:
                    with app.app_context():
                        emit_progress(
                            int(((processed_files + 1) * 100) / total_files),
                            f'文件 {filename} 处理失败：没有有效数据'
                        )
                    all_dataframes[sheet_name] = pl.DataFrame({"message": ["没有有效数据"]})

            except Exception as e:
                logger.error(f"Error processing file {filename}: {str(e)}")
                with app.app_context():
                    emit_progress(
                        int(((processed_files + 1) * 100) / total_files),
                        f'文件 {filename} 处理出错: {str(e)}'
                    )
                error_sheet_name = sheet_name if 'sheet_name' in locals() else f"error_{processed_files}"
                all_dataframes[error_sheet_name[:31]] = pl.DataFrame({"message": [f"处理出错: {str(e)}"]})

            finally:
                processed_files += 1

        # 创建一个新的Excel工作簿
        wb = Workbook()
        # 删除默认创建的sheet
        wb.remove(wb.active)
        
        # 将所有数据框写入Excel文件
        for sheet_name, df in all_dataframes.items():
            # 创建新的worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # 写入表头
            headers = df.columns
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # 写入数据
            data = df.to_numpy()
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 应用格式化
            format_worksheet(ws)

        # 保存到BytesIO
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        # 上传到R2
        r2_key = f'{session_id}/{output_filename}'
        s3_client.put_object(
            Bucket=R2_CONFIG['bucket_name'],
            Key=r2_key,
            Body=output_buffer.getvalue()
        )

        with app.app_context():
            emit_progress(100, '所有文件处理完成！')
        return jsonify({'success': r2_key})

    except Exception as e:
        logger.error(f"Error processing files: {str(e)}", exc_info=True)
        with app.app_context():
            emit_progress(100, f'处理过程中发生错误: {str(e)}')
        return jsonify({'error': f'处理文件时发生错误: {str(e)}'})

@app.route('/download/<session_id>/<filename>')
def download_file(session_id, filename):
    try:
        r2_key = f'{session_id}/{filename}'
        
        # 从R2获取文件
        response = s3_client.get_object(
            Bucket=R2_CONFIG['bucket_name'],
            Key=r2_key
        )
        
        file_data = response['Body'].read()
        
        # 创建内存文件流
        file_stream = BytesIO(file_data)
        
        return_data = send_file(
            file_stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # 下载完成后删除R2中的文件
        @return_data.call_on_close
        def cleanup():
            try:
                s3_client.delete_object(
                    Bucket=R2_CONFIG['bucket_name'],
                    Key=r2_key
                )
            except Exception as e:
                logger.warning(f"Failed to cleanup R2 object: {e}")

        return return_data

    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}", exc_info=True)
        return jsonify({'error': f'下载文件时发生错误: {str(e)}'})

@app.errorhandler(Exception)
def handle_error(error):
    logger.error(f"Unhandled error: {str(error)}", exc_info=True)
    return jsonify({
        "error": "Internal server error",
        "message": str(error)
    }), 500

if __name__ == '__main__':
    # 本地开发时使用
    app.run(host='0.0.0.0', port=5678)
else:
    # Vercel部署时使用
    app = app